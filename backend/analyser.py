"""
Bank Statement Analyser
Supports HDFC, CUB, IOB, PNB, SBI text-based PDF statements.

Parses a bank statement PDF, enriches each transaction (merchant + spending
category), runs simple analytics (monthly/category summaries, top merchants,
anomaly detection) and writes a formatted multi-sheet Excel report.

Bank is auto-detected from the PDF header. Each bank has its own extraction
strategy (word-based spatial columns or table-based).

CLI:
    python Bank_Statement_Analyser.py [input.pdf] [output.xlsx] [stats.json]

Programmatic (used by the Flask server):
    from Bank_Statement_Analyser import analyse
    stats = analyse("input.pdf", "output.xlsx")
"""

import csv
import os
import re
import sys
import pdfplumber
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────

DEFAULT_INPUT_PDF   = "Account Statement.pdf"
DEFAULT_OUTPUT_XLSX = "Bank_Statement_Report.xlsx"
ANOMALY_Z = 2.0

# Column x-boundaries discovered from HDFC PDF structure.
# Each tuple: (col_name, x_start, x_end)
# Words whose x0 falls in [x_start, x_end) are assigned to that column.
HDFC_COLS = [
    ("date",       28,  60),
    ("narration",  60, 280),
    ("ref",       280, 360),
    ("value_date",360, 405),
    ("debit",     405, 490),
    ("credit",    490, 560),
    ("balance",   560, 700),
]

# Category keyword rules. Single source of truth shared with the dbt
# `category_keywords` seed, so the Python and SQL categorisation stay in sync.
# Each rule: (priority, category, keyword, requires_credit). Lowest priority
# wins; requires_credit encodes the salary credit-gate.
CATEGORY_KEYWORDS_CSV = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "category_keywords.csv",
)


def load_category_rules(path: str = CATEGORY_KEYWORDS_CSV) -> list[tuple[int, str, str, bool]]:
    """Load the shared category keyword seed, sorted by ascending priority."""
    rules = []
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rules.append((
                int(row["priority"]),
                row["category"],
                row["keyword"].lower(),
                row["requires_credit"].strip().lower() == "true",
            ))
    rules.sort(key=lambda r: r[0])
    return rules


CATEGORY_RULES = load_category_rules()

# ── Date / amount helpers ───────────────────────────────────────────────────

DATE_FMTS = ["%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y",
             "%d %b %Y", "%d %b %y", "%d-%b-%Y", "%d-%b-%y"]

def parse_date(s: str):
    s = str(s).strip()
    for fmt in DATE_FMTS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def is_date(s: str) -> bool:
    return parse_date(s) is not None

def parse_amount(s) -> float:
    if not s:
        return 0.0
    cleaned = re.sub(r"[^\d.]", "", str(s))
    try:
        return float(cleaned) if cleaned else 0.0
    except ValueError:
        return 0.0

# ── Core extraction ───────────────────────────────────────────────────────────

def assign_col(x0: float) -> str | None:
    for name, x_start, x_end in HDFC_COLS:
        if x_start <= x0 < x_end:
            return name
    return None

# Markers that only appear in the HDFC STATEMENT SUMMARY footer block.
# Deliberately NOT including "OPENINGBALANCE": that label can also be a
# legitimate opening-balance row at the TOP of some statements, and stopping
# there would drop every transaction.
SUMMARY_MARKERS = ("STATEMENTSUMMARY", "DRCOUNT", "CRCOUNT", "CLOSINGBAL", "GENERATEDON")

def _is_summary_line(line_words) -> bool:
    """True once we reach the STATEMENT SUMMARY / totals / 'Generated On' block."""
    joined = "".join(w["text"] for w in line_words).upper()
    return any(marker in joined for marker in SUMMARY_MARKERS)

def _parse_page_words(words) -> tuple[list[dict], bool]:
    """
    Parse one page's words into transaction rows.

    Returns (rows, reached_summary). When reached_summary is True the caller
    should stop processing further pages — the summary block ends the table.

    Pulled out of extract_transactions so the row-assembly logic (and the
    summary-boundary handling) is unit-testable without a real PDF.
    """
    # Find the header row y — or fall back to the table start zone.
    # Page 2+ may not have a visible header (continuation page).
    header_top = None
    for w in words:
        if w["text"] in ("Date", "Narration"):
            header_top = w["top"]
            break

    # If no header found, look for the first date-like token below the
    # address block (address ends by ~y=210) and use that as the start.
    if header_top is None:
        for w in sorted(words, key=lambda x: x["top"]):
            if w["top"] > 215 and is_date(w["text"]) and w["x0"] < 50:
                header_top = w["top"] - 5  # start just above first row
                break
    if header_top is None:
        return [], False  # truly no table on this page

    # Group words into lines by quantised y (±3px tolerance)
    lines: dict[int, list] = {}
    for w in words:
        if w["top"] <= header_top:
            continue
        y_key = round(w["top"] / 8.6) * 8   # HDFC row height ≈ 17.2pt → bucket by half
        lines.setdefault(y_key, []).append(w)

    rows: list[dict] = []
    pending: dict | None = None
    reached_summary = False
    skip_tail = False

    for y_key in sorted(lines):
        line_words = sorted(lines[y_key], key=lambda w: w["x0"])

        # *** FIX *** Once this page hits the STATEMENT SUMMARY / totals / footer
        # block, ignore the rest of THIS page so the grand-total figures can't be
        # absorbed as a continuation line. We do NOT stop the whole document —
        # later pages still get parsed (statements run to many pages).
        if skip_tail:
            continue
        if _is_summary_line(line_words):
            reached_summary = True
            skip_tail = True
            continue

        # Build column buckets for this line
        buckets: dict[str, list[str]] = {n: [] for n, *_ in HDFC_COLS}
        for w in line_words:
            col = assign_col(w["x0"])
            if col:
                buckets[col].append(w["text"])

        date_str = " ".join(buckets["date"])
        narr_str = " ".join(buckets["narration"])
        ref_str  = " ".join(buckets["ref"])
        vdt_str  = " ".join(buckets["value_date"])
        dbt_str  = " ".join(buckets["debit"])
        crd_str  = " ".join(buckets["credit"])
        bal_str  = " ".join(buckets["balance"])

        if is_date(date_str):
            # New transaction row — save previous
            if pending:
                rows.append(pending)
            pending = {
                "date":       parse_date(date_str),
                "narration":  narr_str,
                "ref_no":     ref_str,
                "value_date": parse_date(vdt_str) or parse_date(date_str),
                "debit":      parse_amount(dbt_str),
                "credit":     parse_amount(crd_str),
                "balance":    parse_amount(bal_str),
            }
        elif pending:
            # Continuation line — append narration, fill only-missing amounts
            if narr_str:
                pending["narration"] += " " + narr_str
            if not pending["debit"]   and dbt_str: pending["debit"]   = parse_amount(dbt_str)
            if not pending["credit"]  and crd_str: pending["credit"]  = parse_amount(crd_str)
            if not pending["balance"] and bal_str: pending["balance"] = parse_amount(bal_str)

    if pending:
        rows.append(pending)
    return rows, reached_summary

# ── Bank detection ────────────────────────────────────────────────────────────

BANK_SIGNATURES = {
    # Name-based (when not redacted). Specific full names come first: UPI
    # narrations mention other banks' short codes (e.g. ".../HDFC/handle"),
    # so the loose "HDFC" fallback must be checked last.
    "CUB":  ["CITY UNION BANK"],
    "IOB":  ["INDIAN OVERSEAS BANK"],
    "PNB":  ["PUNJAB NATIONAL BANK"],
    # "STATE BANK OF INDIA" often isn't present as extractable text at all on
    # the "Account Summary" layout (the header/logo area renders as an image,
    # not real text) — "DRAWING POWER"/"UNCLEARED AMOUNT" are SBI-specific
    # passbook field labels that reliably appear in that layout's account
    # summary block instead, and must be listed before "HDFC" below since a
    # UPI counterparty mention of "HDFC" in a transaction narration would
    # otherwise false-positive-match first.
    "SBI":  ["STATE BANK OF INDIA", "DRAWING POWER", "UNCLEARED AMOUNT"],
    # Specific phrase, not just "AXIS BANK" — that string also appears as a
    # UPI counterparty bank name in other banks' statements (false-positive risk).
    "AXIS": ["STATEMENT OF AXIS ACCOUNT"],
    "HDFC": ["HDFC BANK", "HDFC"],
    # Format/structural markers (work even when bank name is redacted)
    "_IOB": ["DEBIT(RS)", "CREDIT(RS)", "DATE(VALUE"],    # IOB column header style
    "_PNB": ["DR AMOUNT", "CR AMOUNT", "TXN NO."],        # PNB column names
    "_SBI": ["(CID:9)", "ACCOUNT STATEMENT FROM"],        # SBI tab-encoding artefact
    "_CUB": ["CITY UNION"],
}

# Markers that end the letterhead/title block and start the transaction
# table — bank-name text on this side of the marker is a reliable signature;
# text after it is transaction narration, which routinely mentions OTHER
# banks (UPI/NEFT counterparties) and would otherwise cause false matches
# (e.g. an Axis statement's UPI narrations naming "State Bank Of India").
_TABLE_START_MARKERS = ("TRAN DATE", "TXN DATE", "VALUE DATE", "OPENING BALANCE")

def _match_bank_signature(text: str) -> str:
    """Pure matching logic over already-uppercased text, split out from
    detect_bank() so it's unit-testable without a real PDF fixture."""
    # Pass 1: header only (before the transaction table) — avoids false
    # matches from counterparty bank names inside transaction narrations.
    cut_points = [text.find(m) for m in _TABLE_START_MARKERS if m in text]
    header_text = text[:min(cut_points)] if cut_points else text
    for bank, sigs in BANK_SIGNATURES.items():
        real_bank = bank.lstrip("_")   # "_IOB" → "IOB"
        if any(sig in header_text for sig in sigs):
            return real_bank

    # Pass 2: fall back to the full scanned text (structural markers, or
    # statements whose bank name doesn't appear before the table).
    for bank, sigs in BANK_SIGNATURES.items():
        real_bank = bank.lstrip("_")
        if any(sig in text for sig in sigs):
            return real_bank
    return "UNKNOWN"


def detect_bank(pdf_path: str) -> str:
    """Identify the issuing bank from name or structural markers across all pages."""
    with pdfplumber.open(pdf_path) as pdf:
        # Scan up to 2 pages for robustness
        text = " ".join(
            (page.extract_text() or "") for page in pdf.pages[:2]
        ).upper()
    return _match_bank_signature(text)

# ── CUB (City Union Bank) — word-based ────────────────────────────────────────
# Each transaction spans 2–3 sub-lines ~5px apart (narration / date+amounts /
# optional continuation). Transactions are ~26px apart, so grouping lines within
# 12px merges intra-transaction sub-lines while keeping transactions separate.

CUB_COLS = [
    ("date",      25,  103),
    ("narration", 103, 265),
    ("cheque",    265, 390),
    ("debit",     390, 470),
    ("credit",    470, 525),
    ("balance",   525, 700),
]

def _extract_cub(pdf) -> list[dict]:
    rows = []
    for page in pdf.pages:
        words = page.extract_words(x_tolerance=3, y_tolerance=3)
        if not words:
            continue
        # Find where the transaction table starts (header row)
        header_top = None
        for w in words:
            if w["text"] == "Date" and w["x0"] < 50:
                header_top = w["top"]
                break
        if header_top is None:
            # Continuation page — start before first date word in date column
            for w in sorted(words, key=lambda x: x["top"]):
                if w["top"] > 50 and is_date(w["text"]) and 25 <= w["x0"] < 103:
                    header_top = w["top"] - 5
                    break
        if header_top is None:
            continue

        # Build raw line dict: y → list of words
        raw_lines: dict[float, list] = {}
        for w in words:
            if w["top"] <= header_top:
                continue
            raw_lines.setdefault(round(w["top"], 1), []).append(w)

        # Merge consecutive sub-lines within 12px into transaction groups
        sorted_ys = sorted(raw_lines)
        groups: list[list[float]] = []
        cur: list[float] = []
        for y in sorted_ys:
            if not cur or y - cur[-1] <= 12:
                cur.append(y)
            else:
                groups.append(cur)
                cur = [y]
        if cur:
            groups.append(cur)

        for grp in groups:
            all_words = []
            for y in grp:
                all_words.extend(raw_lines[y])

            buckets: dict[str, list[str]] = {n: [] for n, *_ in CUB_COLS}
            for w in sorted(all_words, key=lambda x: (x["top"], x["x0"])):
                for name, x0, x1 in CUB_COLS:
                    if x0 <= w["x0"] < x1:
                        buckets[name].append(w["text"])
                        break

            date_str = " ".join(buckets["date"])
            narr_str = " ".join(buckets["narration"])
            dbt_str  = " ".join(buckets["debit"])
            crd_str  = " ".join(buckets["credit"])
            bal_str  = " ".join(buckets["balance"])

            date = parse_date(date_str)
            if not date:
                continue
            rows.append({
                "date":       date,
                "narration":  narr_str,
                "ref_no":     " ".join(buckets["cheque"]).replace("-", "").strip(),
                "value_date": date,
                "debit":      parse_amount(dbt_str),
                "credit":     parse_amount(crd_str),
                "balance":    parse_amount(bal_str),
            })
    return rows

# ── IOB (Indian Overseas Bank) — word-based ────────────────────────────────────
# Standard row-per-line layout. Debit and credit columns identified by x-range;
# "-" placeholder in the empty column parses as 0.0 automatically.

IOB_COLS = [
    ("date",      44, 112),
    ("narration", 112, 278),
    ("ref",       278, 343),
    ("type",      343, 408),
    ("debit",     408, 462),
    ("credit",    462, 510),
    ("balance",   510, 700),
]

IOB_STOP_MARKERS = ("AVAILABLEBALANCE", "COMPUTERGENERATEDSTATEMENT", "DOESNOTREQUIRE")

def _extract_iob(pdf) -> list[dict]:
    rows = []
    for page in pdf.pages:
        words = page.extract_words(x_tolerance=3, y_tolerance=3)
        if not words:
            continue
        # Find header — IOB uses "Date(Value" (not plain "Date") at x≈49
        header_top = None
        for w in words:
            if "Date" in w["text"] and w["x0"] < 60 and w["top"] > 200:
                header_top = w["top"]
                break
        if header_top is None:
            continue

        raw_lines: dict[float, list] = {}
        for w in words:
            if w["top"] <= header_top:
                continue
            raw_lines.setdefault(round(w["top"], 1), []).append(w)

        # IOB: intra-transaction y-span ≈10px, inter-transaction gap ≈14px
        # Merge sub-lines within 10px so date+amounts land in the same group
        sorted_ys = sorted(raw_lines)
        groups: list[list[float]] = []
        cur: list[float] = []
        for y in sorted_ys:
            if not cur or y - cur[-1] <= 10:
                cur.append(y)
            else:
                groups.append(cur)
                cur = [y]
        if cur:
            groups.append(cur)

        for grp in groups:
            all_words = []
            for y in grp:
                all_words.extend(raw_lines[y])

            joined = "".join(w["text"] for w in all_words).upper()
            if any(m in joined for m in IOB_STOP_MARKERS):
                break

            buckets: dict[str, list[str]] = {n: [] for n, *_ in IOB_COLS}
            for w in sorted(all_words, key=lambda x: (x["top"], x["x0"])):
                for name, x0, x1 in IOB_COLS:
                    if x0 <= w["x0"] < x1:
                        buckets[name].append(w["text"])
                        break

            # First parseable date wins (echo "(01-Nov-25)" comes second, is skipped)
            date = None
            for tok in buckets["date"]:
                date = parse_date(tok.strip("()"))
                if date:
                    break
            if not date:
                continue

            narr_str = " ".join(t for t in buckets["narration"] if t)
            dbt_str  = " ".join(d for d in buckets["debit"]  if d != "-")
            crd_str  = " ".join(c for c in buckets["credit"] if c != "-")
            bal_str  = " ".join(buckets["balance"])

            rows.append({
                "date":       date,
                "narration":  narr_str,
                "ref_no":     " ".join(buckets["ref"]),
                "value_date": date,
                "debit":      parse_amount(dbt_str),
                "credit":     parse_amount(crd_str),
                "balance":    parse_amount(bal_str),
            })
    return rows

# ── PNB (Punjab National Bank) — table-based ──────────────────────────────────
# pdfplumber table extraction works cleanly for PNB. Balance field has a
# "Cr." / "Dr." suffix and may contain a newline before the suffix.

def _extract_pnb(pdf) -> list[dict]:
    rows = []
    header_seen = False
    for page in pdf.pages:
        tables = page.extract_tables()
        for table in tables:
            for row in table:
                if not row or len(row) < 7:
                    continue
                # Skip the header row itself
                if row[0] and "Txn" in str(row[0]) and "No" in str(row[0]):
                    header_seen = True
                    continue
                txn_no, txn_date, desc, _branch, _cheque, dr_raw, cr_raw, bal_raw = (
                    (row + [""] * 8)[:8]
                )
                date = parse_date((txn_date or "").strip())
                if not date:
                    continue
                # Balance: remove "Cr." / "Dr." suffix and embedded newlines
                bal_clean = re.sub(r"[^\d,.]", "", (bal_raw or "").replace("\n", " ")).rstrip(".")
                rows.append({
                    "date":       date,
                    "narration":  (desc or "").replace("\n", " ").strip(),
                    "ref_no":     (txn_no or "").strip(),
                    "value_date": date,
                    "debit":      parse_amount(dr_raw),
                    "credit":     parse_amount(cr_raw),
                    "balance":    parse_amount(bal_clean),
                })
    _ = header_seen  # suppress unused-variable warning
    return rows

# ── SBI (State Bank of India) — word-based ────────────────────────────────────
# SBI splits date across two y-positions when the day is two digits
# ("15 May" on line 1, "2025" on line 2, 9px lower). Grouping lines within
# 14px merges the year with its date while keeping separate transactions apart
# (inter-transaction gap ~31px).

SBI_COLS = [
    ("txn_date",  30,  95),
    ("val_date",  95, 143),
    ("narration", 143, 275),
    ("ref",       275, 379),
    ("debit",     379, 448),
    ("credit",    448, 509),
    ("balance",   509, 700),
]

SBI_STOP_MARKERS = ("OPENINGBALANCE", "CLOSINGBALANCE", "TOTALDEBIT", "TOTALCREDIT")

def _extract_sbi(pdf) -> list[dict]:
    rows = []
    for page in pdf.pages:
        words = page.extract_words(x_tolerance=3, y_tolerance=3)
        if not words:
            continue
        # Locate header: look for "Txn" and "Date" close together
        header_top = None
        for w in words:
            if w["text"] == "Txn" and w["x0"] < 60:
                header_top = w["top"]
                break
        if header_top is None:
            # Continuation page: start above first date candidate in txn_date column
            for w in sorted(words, key=lambda x: x["top"]):
                if w["top"] > 30 and w["x0"] < 95 and re.match(r"^\d{1,2}$", w["text"]):
                    header_top = w["top"] - 5
                    break
        if header_top is None:
            continue

        raw_lines: dict[float, list] = {}
        for w in words:
            if w["top"] <= header_top:
                continue
            raw_lines.setdefault(round(w["top"], 1), []).append(w)

        # Merge sub-lines within 11px: intra-row gap=9px ≤11, inter-row gap=13px >11
        sorted_ys = sorted(raw_lines)
        groups: list[list[float]] = []
        cur: list[float] = []
        for y in sorted_ys:
            if not cur or y - cur[-1] <= 11:
                cur.append(y)
            else:
                groups.append(cur)
                cur = [y]
        if cur:
            groups.append(cur)

        pending = None
        for grp in groups:
            all_words = []
            for y in grp:
                all_words.extend(raw_lines[y])

            joined = "".join(w["text"] for w in all_words).upper()
            if any(m in joined for m in SBI_STOP_MARKERS):
                break

            buckets: dict[str, list[str]] = {n: [] for n, *_ in SBI_COLS}
            for w in sorted(all_words, key=lambda x: (x["top"], x["x0"])):
                for name, x0, x1 in SBI_COLS:
                    if x0 <= w["x0"] < x1:
                        buckets[name].append(w["text"])
                        break

            date_str = " ".join(buckets["txn_date"])
            vdt_str  = " ".join(buckets["val_date"])
            narr_str = " ".join(buckets["narration"])
            dbt_str  = " ".join(buckets["debit"])
            crd_str  = " ".join(buckets["credit"])
            bal_str  = " ".join(buckets["balance"])

            date = parse_date(date_str)
            if date:
                if pending:
                    rows.append(pending)
                pending = {
                    "date":       date,
                    "narration":  narr_str,
                    "ref_no":     " ".join(buckets["ref"]),
                    "value_date": parse_date(vdt_str) or date,
                    "debit":      parse_amount(dbt_str),
                    "credit":     parse_amount(crd_str),
                    "balance":    parse_amount(bal_str),
                }
            elif pending:
                if narr_str:
                    pending["narration"] += " " + narr_str
                if not pending["balance"] and bal_str:
                    pending["balance"] = parse_amount(bal_str)

        if pending:
            rows.append(pending)
    return rows

# ── SBI v2 ("STATEMENT OF ACCOUNT" layout) — table-based ─────────────────────
# Newer SBI e-statements use a bordered 7-column table:
# [txn_date, value_date, description(multi-line), cheque_no, debit, credit, balance]
# Empty amount cells contain "-".

def _extract_sbi_v2(pdf) -> list[dict]:
    rows: list[dict] = []
    for page in pdf.pages:
        for table in page.extract_tables():
            for r in table:
                if not r or len(r) < 7:
                    continue
                d = parse_date((r[0] or "").strip())
                if d is None:
                    continue
                narration = " ".join((r[2] or "").split())
                # Strip the leading WDL TFR / DEP TFR transfer marker
                narration = re.sub(r"^(WDL|DEP)\s+TFR\s*", "", narration, flags=re.I)
                blank = lambda cell: not cell or cell.strip() in ("", "-")
                rows.append({
                    "date": d,
                    "narration": narration,
                    "ref": (r[3] or "").strip().strip("-"),
                    "debit": 0.0 if blank(r[4]) else parse_amount(r[4]),
                    "credit": 0.0 if blank(r[5]) else parse_amount(r[5]),
                    "balance": parse_amount(r[6]),
                })
    return rows

# ── Axis Bank — word-based spatial ───────────────────────────────────────────
# Columns: Tran Date | Chq No | Particulars | Debit | Credit | Balance | Init. Br
# Unlike HDFC/SBI, the date+amounts sit on the LAST physical line of a
# transaction block, with narration wrapping across the line(s) above it.

AXIS_COLS = [
    ("date",       25,  95),
    ("chq",        95, 130),
    ("narration", 130, 340),
    ("debit",     340, 400),
    ("credit",    400, 465),
    ("balance",   465, 530),
]
AXIS_DATE_RE = re.compile(r"^\d{2}-\d{2}-\d{4}$")
AXIS_STOP_MARKERS = ("TRANSACTION TOTAL", "CLOSING BALANCE", "END OF STATEMENT", "LEGENDS")

def _assign_axis_col(x0: float) -> str | None:
    for name, x_start, x_end in AXIS_COLS:
        if x_start <= x0 < x_end:
            return name
    return None

def _extract_axis(pdf) -> list[dict]:
    rows: list[dict] = []
    for page in pdf.pages:
        words = page.extract_words(x_tolerance=2, y_tolerance=3)
        if not words:
            continue
        lines: dict[float, list] = {}
        for w in words:
            lines.setdefault(round(w["top"], 1), []).append(w)

        pending_narration: list[str] = []
        for y in sorted(lines):
            line_words = sorted(lines[y], key=lambda w: w["x0"])
            joined = " ".join(w["text"] for w in line_words).upper()
            if any(marker in joined for marker in AXIS_STOP_MARKERS):
                break

            buckets: dict[str, list[str]] = {n: [] for n, *_ in AXIS_COLS}
            for w in line_words:
                col = _assign_axis_col(w["x0"])
                if col:
                    buckets[col].append(w["text"])

            date_str = "".join(buckets["date"])
            if AXIS_DATE_RE.match(date_str):
                narration = " ".join(pending_narration + buckets["narration"]).strip()
                rows.append({
                    "date": parse_date(date_str),
                    "narration": narration,
                    "debit": parse_amount(" ".join(buckets["debit"])),
                    "credit": parse_amount(" ".join(buckets["credit"])),
                    "balance": parse_amount(" ".join(buckets["balance"])),
                })
                pending_narration = []
            elif "OPENING" in joined and "BALANCE" in joined:
                pending_narration = []  # opening-balance line has no date; not a transaction
            elif buckets["narration"]:
                pending_narration.append(" ".join(buckets["narration"]))
    return rows

# ── Dispatcher ────────────────────────────────────────────────────────────────

def extract_transactions(pdf_path: str) -> list[dict]:
    """
    Auto-detect the bank and extract transactions using the appropriate parser.
    Supports HDFC (word/spatial), CUB (word/spatial), IOB (word/spatial),
    PNB (table-based), SBI (word/spatial).
    """
    bank = detect_bank(pdf_path)
    with pdfplumber.open(pdf_path) as pdf:
        if bank == "CUB":
            return _extract_cub(pdf)
        if bank == "IOB":
            return _extract_iob(pdf)
        if bank == "PNB":
            return _extract_pnb(pdf)
        if bank == "SBI":
            # Two SBI layouts in the wild: the older tab-encoded spatial one
            # and the newer "STATEMENT OF ACCOUNT" bordered table. Try the
            # old parser first, fall back to the table parser.
            return _extract_sbi(pdf) or _extract_sbi_v2(pdf)
        if bank == "AXIS":
            return _extract_axis(pdf)
        # Default: HDFC (or UNKNOWN — try HDFC spatial layout)
        all_rows: list[dict] = []
        for page in pdf.pages:
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            if not words:
                continue
            page_rows, _ = _parse_page_words(words)
            all_rows.extend(page_rows)
        return all_rows

# ── Enrichment ────────────────────────────────────────────────────────────────

def extract_merchant(narration: str) -> str:
    """Pull a clean merchant name from UPI / POS / NEFT narration strings."""
    text = narration.strip()

    # UPI pattern: UPI-MERCHANTNAME-...
    m = re.match(r"UPI-([A-Za-z0-9 &']+?)(?:-[A-Z0-9@]{4,}|-\d|$)", text, re.IGNORECASE)
    if m:
        return m.group(1).strip().title()[:40]

    # SBI slash style: UPI/DR/609178413960/SALEEM K/YESB/handle/...
    m = re.match(r"UPI/(?:DR|CR)/\d+/([A-Za-z][A-Za-z0-9 .&']*)", text, re.IGNORECASE)
    if m:
        return m.group(1).strip().title()[:40]

    # Axis slash style: UPI/P2A/646193132733/XAVIER INFANTA MICHE /UPI/BANK
    # or UPI/P2M/125430393560/Swiggy /NO REM/ICICI Bank
    m = re.match(r"UPI/P2[AM]/\d+/([A-Za-z][A-Za-z0-9 .&']*?)\s*/", text, re.IGNORECASE)
    if m:
        return m.group(1).strip().title()[:40]

    # POS pattern
    m = re.search(r"POS\s+([A-Za-z][A-Za-z0-9 &']+)", text, re.IGNORECASE)
    if m:
        return m.group(1).strip().title()[:40]

    # NEFT/IMPS/RTGS: NEFT/refcode/NAME/... — non-greedy so a following
    # bank/branch segment (also alphabetic) doesn't get swallowed too.
    m = re.match(r"(?:NEFT|IMPS|RTGS)/[A-Z0-9]+/([A-Za-z][A-Za-z0-9 .&']*?)\s*/", text, re.IGNORECASE)
    if m:
        return m.group(1).strip().title()[:40]

    # Axis mobile transfer: MOB/TPFT/NAME/account_number
    m = re.match(r"MOB/TPFT/([A-Za-z][A-Za-z0-9 .&']*?)\s*/", text, re.IGNORECASE)
    if m:
        return m.group(1).strip().title()[:40]

    return text[:40]

def assign_category(narration: str, credit: float) -> str:
    text = narration.lower()
    for _priority, category, keyword, requires_credit in CATEGORY_RULES:
        if requires_credit and not credit > 0:
            continue
        if keyword in text:
            return category
    return "Other Income" if credit > 0 else "Other Expense"

def clean_and_enrich(raw: list[dict]) -> list[dict]:
    seen = set()
    out = []
    for r in raw:
        if r.get("date") is None:
            continue
        key = (r["date"], r["narration"][:30], r["debit"], r["credit"])
        if key in seen:
            continue
        seen.add(key)
        r["merchant"] = extract_merchant(r["narration"])
        r["category"] = assign_category(r["narration"], r["credit"])
        r["is_anomaly"] = False  # set later by detect_anomalies / analyse
        out.append(r)
    out.sort(key=lambda x: x["date"])
    return out

# ── Analytics ─────────────────────────────────────────────────────────────────

def _debit_stats(rows):
    """Shared debit statistics used by anomaly detection and the report."""
    debits = [r["debit"] for r in rows if r["debit"] > 0]
    if not debits:
        return {"debits": [], "mean": 0.0, "std": 0.0, "thresh": 0.0}
    mean = sum(debits) / len(debits)
    std  = (sum((x - mean) ** 2 for x in debits) / len(debits)) ** 0.5
    return {"debits": debits, "mean": mean, "std": std, "thresh": mean + ANOMALY_Z * std}

def monthly_summary(rows):
    months = {}
    for r in rows:
        key = r["date"].strftime("%b %Y")
        months.setdefault(key, {"month": key, "income": 0.0, "expense": 0.0})
        months[key]["income"]  += r["credit"]
        months[key]["expense"] += r["debit"]
    for m in months.values():
        m["net"] = m["income"] - m["expense"]
    return list(months.values())

def category_summary(rows):
    cats = {}
    for r in rows:
        cat = r["category"]
        cats.setdefault(cat, {"category": cat, "spend": 0.0, "txn_count": 0})
        cats[cat]["spend"]     += r["debit"] if r["debit"] else r["credit"]
        cats[cat]["txn_count"] += 1
    return sorted(cats.values(), key=lambda x: x["spend"], reverse=True)

def top_merchants(rows):
    m = {}
    for r in rows:
        if r["debit"] > 0:
            m[r["merchant"]] = m.get(r["merchant"], 0) + r["debit"]
    return [{"merchant": k, "total_spend": v}
            for k, v in sorted(m.items(), key=lambda x: x[1], reverse=True)[:10]]

def detect_anomalies(rows):
    stats = _debit_stats(rows)
    if len(stats["debits"]) < 3:
        return []
    thresh = stats["thresh"]
    return sorted([r for r in rows if r["debit"] > thresh],
                  key=lambda x: x["debit"], reverse=True)

def spending_stats(rows):
    debits  = [r["debit"]  for r in rows if r["debit"]  > 0]
    credits = [r["credit"] for r in rows if r["credit"] > 0]
    return {
        "total_spend":     sum(debits),
        "total_income":    sum(credits),
        "net_cash_flow":   sum(credits) - sum(debits),
        "avg_expense":     sum(debits) / len(debits) if debits else 0,
        "largest_expense": max(debits)  if debits  else 0,
        "largest_credit":  max(credits) if credits else 0,
        "txn_count":       len(rows),
        "expense_count":   len(debits),
        "income_count":    len(credits),
    }

# ── Excel export ──────────────────────────────────────────────────────────────

C_NAVY  = "1A3C5E"
C_WHITE = "FFFFFF"
C_ALT   = "EFF4F9"
C_RED   = "C0392B"
C_GREEN = "1E8449"
C_BLUE  = "2471A3"
C_WARN  = "FDECEA"
C_WARN_T= "922B21"

THIN   = Side(style="thin", color="CCCCCC")
BDR    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
INR    = '₹#,##0.00'
DTEFMT = 'DD/MM/YYYY'

def hdr(ws, row, col, value, width=None, fill=C_NAVY):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=True, color=C_WHITE, size=10)
    c.fill      = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = BDR
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def autofit_col(ws, col, min_width=10, max_width=60):
    """Widen a column to fit its longest value, capped so one long UPI
    narration or merchant name can't blow out the whole sheet layout."""
    letter = get_column_letter(col)
    max_len = max((len(str(c.value)) for c in ws[letter] if c.value is not None), default=0)
    ws.column_dimensions[letter].width = max(min(max_len + 3, max_width), min_width)

def dc(ws, row, col, value, fmt=None, bold=False, color=None, bg=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=9, bold=bold, color=color or "000000")
    c.border    = BDR
    c.alignment = Alignment(vertical="center")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    elif row % 2 == 0:
        c.fill = PatternFill("solid", fgColor=C_ALT)
    if fmt:
        c.number_format = fmt
    return c

def write_summary(wb, stats):
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20

    t = ws.cell(row=1, column=1, value="Bank Statement — Analysis")
    t.font = Font(name="Arial", bold=True, size=14, color=C_BLUE)
    ws.merge_cells("A1:B1")

    pairs = [
        ("Total Income",         stats["total_income"],    INR,  C_GREEN),
        ("Total Expenses",       stats["total_spend"],     INR,  C_RED),
        ("Net Cash Flow",        stats["net_cash_flow"],   INR,  C_GREEN if stats["net_cash_flow"] >= 0 else C_RED),
        ("Total Transactions",   stats["txn_count"],       None, C_BLUE),
        ("Expense Transactions", stats["expense_count"],   None, C_RED),
        ("Income Transactions",  stats["income_count"],    None, C_GREEN),
        ("Avg Expense",          stats["avg_expense"],     INR,  C_RED),
        ("Largest Expense",      stats["largest_expense"], INR,  C_RED),
        ("Largest Credit",       stats["largest_credit"],  INR,  C_GREEN),
    ]
    for i, (label, val, fmt, color) in enumerate(pairs, 3):
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = Font(name="Arial", size=10, bold=True)
        vc = ws.cell(row=i, column=2, value=val)
        vc.font = Font(name="Arial", size=10, bold=(label == "Net Cash Flow"), color=color)
        if fmt:
            vc.number_format = fmt

def write_transactions(wb, rows):
    ws = wb.create_sheet("Transactions")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    cols = [("Date",12),("Merchant",24),("Narration",46),("Ref No",20),
            ("Value Date",12),("Debit (₹)",14),("Credit (₹)",14),
            ("Balance (₹)",14),("Category",22),("Flag",8)]
    for i, (h, w) in enumerate(cols, 1):
        hdr(ws, 1, i, h, w)

    for idx, r in enumerate(rows, 2):
        is_anom = r.get("is_anomaly", False)
        bg = C_WARN if is_anom else None
        dc(ws, idx, 1, r["date"],       DTEFMT, bg=bg)
        dc(ws, idx, 2, r["merchant"],           bg=bg)
        dc(ws, idx, 3, r["narration"],          bg=bg)
        dc(ws, idx, 4, r.get("ref_no", ""),                    bg=bg)
        dc(ws, idx, 5, r.get("value_date") or r["date"], DTEFMT, bg=bg)
        dc(ws, idx, 6, r["debit"]  or None, INR, color=C_RED   if r["debit"]  > 0 else None, bg=bg)
        dc(ws, idx, 7, r["credit"] or None, INR, color=C_GREEN if r["credit"] > 0 else None, bg=bg)
        dc(ws, idx, 8, r["balance"],        INR, bg=bg)
        dc(ws, idx, 9, r["category"],           bg=bg)
        dc(ws, idx,10, "⚠" if is_anom else "",
           bold=is_anom, color=C_WARN_T if is_anom else None, bg=bg)

    ws.auto_filter.ref = "A1:J1"
    autofit_col(ws, 2)  # Merchant
    autofit_col(ws, 3)  # Narration

def write_monthly(wb, monthly):
    ws = wb.create_sheet("Monthly Summary")
    for i, (h, w) in enumerate([("Month",12),("Income (₹)",16),("Expense (₹)",16),("Net (₹)",16)], 1):
        hdr(ws, 1, i, h, w)
    for idx, m in enumerate(monthly, 2):
        dc(ws, idx, 1, m["month"])
        dc(ws, idx, 2, m["income"],  INR, color=C_GREEN)
        dc(ws, idx, 3, m["expense"], INR, color=C_RED)
        dc(ws, idx, 4, m["net"], INR, bold=True, color=C_GREEN if m["net"] >= 0 else C_RED)

    if monthly:
        from openpyxl.chart.label import DataLabelList
        last = len(monthly) + 1
        chart = BarChart()
        chart.type, chart.grouping = "col", "clustered"
        chart.title = "Monthly Income vs Expense"
        chart.y_axis.title  = "Amount (Rs.)"
        chart.y_axis.numFmt = "#,##0"
        chart.width, chart.height = 26, 14
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=last), titles_from_data=True)
        chart.add_data(Reference(ws, min_col=3, min_row=1, max_row=last), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))
        chart.series[0].graphicalProperties.solidFill = "1E8449"
        chart.series[1].graphicalProperties.solidFill = "C0392B"
        for s in chart.series:
            dlbls = DataLabelList()
            dlbls.showVal = True
            dlbls.showLegendKey = False
            dlbls.showCatName   = False
            dlbls.showSerName   = False
            dlbls.showPercent   = False
            s.dLbls = dlbls
        ws.add_chart(chart, "F2")

def write_categories(wb, cats):
    ws = wb.create_sheet("Categories")
    for i, (h, w) in enumerate([("Category",24),("Total (₹)",16),("Transactions",14)], 1):
        hdr(ws, 1, i, h, w)
    for idx, c in enumerate(cats, 2):
        dc(ws, idx, 1, c["category"])
        dc(ws, idx, 2, c["spend"], INR)
        dc(ws, idx, 3, c["txn_count"])

    if cats:
        last = len(cats) + 1
        pie = PieChart()
        pie.title = "Spending by Category"
        pie.width, pie.height = 18, 12
        pie.add_data(Reference(ws, min_col=2, min_row=1, max_row=last), titles_from_data=True)
        pie.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))
        ws.add_chart(pie, "E2")

def write_merchants(wb, merchants):
    ws = wb.create_sheet("Top Merchants")
    hdr(ws, 1, 1, "Merchant",           28)
    hdr(ws, 1, 2, "Total Spend (Rs.)",  18)
    for idx, m in enumerate(merchants, 2):
        dc(ws, idx, 1, m["merchant"])
        dc(ws, idx, 2, m["total_spend"], INR)

    if merchants:
        from openpyxl.chart.label import DataLabelList
        last = len(merchants) + 1
        bar = BarChart()
        bar.type  = "bar"
        bar.title = "Top Merchants by Spend"
        bar.x_axis.title = "Amount Spent (Rs.)"
        bar.x_axis.numFmt = "#,##0"
        bar.x_axis.majorGridlines = None
        bar.width, bar.height = 28, 18
        bar.add_data(Reference(ws, min_col=2, min_row=1, max_row=last), titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))
        bar.series[0].graphicalProperties.solidFill = C_BLUE
        dlbls = DataLabelList()
        dlbls.showVal = True
        dlbls.showLegendKey = False
        dlbls.showCatName   = False
        dlbls.showSerName   = False
        dlbls.showPercent   = False
        bar.series[0].dLbls = dlbls
        ws.add_chart(bar, "D2")

def write_anomalies(wb, anomalies, all_rows):
    ws = wb.create_sheet("Anomalies")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 44
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 22

    # ── Plain-English explanation block ──────────────────────────────────────
    s = _debit_stats(all_rows)
    mean, thresh = s["mean"], s["thresh"]
    n_debits = len(s["debits"])

    title_cell = ws.cell(row=1, column=1,
        value="Unusual Transactions — These payments are much larger than normal (click to jump to all transactions)")
    title_cell.font = Font(name="Arial", bold=True, size=12, color=C_WARN_T, underline="single")
    title_cell.hyperlink = "#'Transactions'!A1"
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 20

    explain_lines = [
        f"How this works: We looked at all {n_debits} expense transactions in your statement.",
        f"Your typical (average) transaction is Rs. {mean:,.0f}.",
        f"Transactions more than Rs. {thresh:,.0f} are flagged below — these stand out as unusually large.",
        f"This helps you spot: accidental double payments, fraud, or big spends you may have forgotten.",
    ]
    for i, line in enumerate(explain_lines, 2):
        c = ws.cell(row=i, column=1, value=line)
        c.font = Font(name="Arial", size=10,
                      color="555555" if i > 2 else "222222",
                      italic=(i > 2))
        ws.merge_cells(f"A{i}:E{i}")
        ws.row_dimensions[i].height = 16

    # Blank separator row
    ws.row_dimensions[6].height = 8

    if not anomalies:
        no_flag = ws.cell(row=7, column=1,
            value="Good news — no unusually large transactions found in this statement.")
        no_flag.font = Font(name="Arial", size=11, color=C_GREEN, bold=True)
        ws.merge_cells("A7:E7")
        return

    # ── Header row ───────────────────────────────────────────────────────────
    for i, (h, w) in enumerate([("Date",12),("Who was paid",28),
                                  ("Full narration",44),
                                  ("Amount paid (Rs.)",18),("Category",22)], 1):
        hdr(ws, 7, i, h, fill=C_WARN_T)

    # ── Data rows ────────────────────────────────────────────────────────────
    for idx, r in enumerate(anomalies, 8):
        dc(ws, idx, 1, r["date"],       DTEFMT, bg=C_WARN)
        dc(ws, idx, 2, r["merchant"],           bg=C_WARN)
        dc(ws, idx, 3, r["narration"],          bg=C_WARN)
        dc(ws, idx, 4, r["debit"], "#,##0.00",
           bold=True, color=C_WARN_T, bg=C_WARN)
        dc(ws, idx, 5, r["category"],           bg=C_WARN)

    # ── Per-row plain-English reason ─────────────────────────────────────────
    reason_row = len(anomalies) + 9
    ws.cell(row=reason_row, column=1,
            value="Why each transaction above was flagged:").font = Font(
                name="Arial", size=10, bold=True, color="444444")
    ws.merge_cells(f"A{reason_row}:E{reason_row}")

    for i, r in enumerate(anomalies, reason_row + 1):
        how_big = r["debit"] / mean if mean else 0
        reason = (
            f"{r['merchant']} — Rs. {r['debit']:,.0f} on {r['date'].strftime('%d %b %Y')}. "
            f"This is {how_big:.1f}x your average spend of Rs. {mean:,.0f}. "
            f"Category: {r['category']}."
        )
        c = ws.cell(row=i, column=1, value=reason)
        c.font  = Font(name="Arial", size=9, color="444444")
        c.fill  = PatternFill("solid", fgColor="FFF8F7")
        ws.merge_cells(f"A{i}:E{i}")
        ws.row_dimensions[i].height = 18

def export_excel(rows, monthly, cats, merchants, anomalies, stats, path):
    wb = Workbook()
    write_summary(wb, stats)
    write_transactions(wb, rows)
    write_monthly(wb, monthly)
    write_categories(wb, cats)
    write_merchants(wb, merchants)
    write_anomalies(wb, anomalies, rows)
    wb.save(path)

# ── Pipeline entry point ────────────────────────────────────────────────────

def analyse(pdf_path: str, output_xlsx: str, stats_path: str | None = None) -> dict:
    """
    Run the full pipeline and return a summary stats dict.

    Single reusable entry point: both the CLI (main) and the Flask server call
    this, so there is one code path instead of shelling out to a script.
    """
    raw = extract_transactions(pdf_path)
    rows = clean_and_enrich(raw)
    if not rows:
        raise ValueError(
            "No transactions found. Is the PDF text-based (not a scanned image)?"
        )

    monthly   = monthly_summary(rows)
    cats      = category_summary(rows)
    merchants = top_merchants(rows)
    anomalies = detect_anomalies(rows)
    for r in anomalies:
        r["is_anomaly"] = True
    stats     = spending_stats(rows)

    export_excel(rows, monthly, cats, merchants, anomalies, stats, output_xlsx)

    result = {
        "total_spend":   round(stats["total_spend"], 2),
        "total_income":  round(stats["total_income"], 2),
        "net_cash_flow": round(stats["net_cash_flow"], 2),
        "transactions":  stats["txn_count"],
        "anomalies":     len(anomalies),
        "top_category":  cats[0]["category"] if cats else "N/A",
    }
    if stats_path:
        import json
        with open(stats_path, "w") as f:
            json.dump(result, f)
    return result

# ── CLI ─────────────────────────────────────────────────────────────────────

def main():
    import os

    pdf        = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_INPUT_PDF
    out        = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT_XLSX
    stats_path = sys.argv[3] if len(sys.argv) > 3 else None

    if not os.path.exists(pdf):
        print(f"ERROR: '{pdf}' not found. Place your PDF here.")
        sys.exit(1)

    print(f"\n[1/2] Analysing '{pdf}'…")
    try:
        result = analyse(pdf, out, stats_path)
    except ValueError as e:
        print(f"  {e}")
        sys.exit(2)

    print(f"[2/2] Wrote '{out}'")
    print(f"\n✅  Done!")
    print(f"    Transactions  : {result['transactions']}")
    print(f"    Total spend   : ₹{result['total_spend']:,.2f}")
    print(f"    Total income  : ₹{result['total_income']:,.2f}")
    print(f"    Net cash flow : ₹{result['net_cash_flow']:,.2f}")
    print(f"    Anomalies     : {result['anomalies']} flagged")

if __name__ == "__main__":
    main()
