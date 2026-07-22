"""SpendStory unit tests — pure-function coverage, no real bank data needed."""
import io
import sys, os
from datetime import date
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

from openpyxl import load_workbook

from analyser import (parse_date, parse_amount, extract_merchant, assign_category,
                      detect_anomalies, monthly_summary, category_summary,
                      top_merchants, spending_stats, clean_and_enrich, export_excel,
                      _match_bank_signature, fy_summary, _mask_text, _mask_ref,
                      _extract_iob)


def row(d, narration="x", debit=0.0, credit=0.0, balance=0.0, merchant="m", category="c"):
    return {"date": d, "narration": narration, "debit": debit, "credit": credit,
            "balance": balance, "merchant": merchant, "category": category, "is_anomaly": False,
            "ref_no": "", "value_date": d}


def test_parse_date_formats():
    assert parse_date("02/04/2026") == date(2026, 4, 2)
    assert parse_date("05-04-2026") == date(2026, 4, 5)
    assert parse_date("2 Apr 2026") == date(2026, 4, 2)
    assert parse_date("not a date") is None
    assert parse_date("") is None


def test_parse_amount():
    assert parse_amount("31,600.00") == 31600.0
    assert parse_amount("₹1,234.56") == 1234.56
    assert parse_amount("-") == 0.0
    assert parse_amount(None) == 0.0
    assert parse_amount("") == 0.0


def test_extract_merchant_patterns():
    # HDFC dash style
    assert extract_merchant("UPI-SWIGGY-SWIGGY@YBL-123") == "Swiggy"
    # SBI slash style
    assert extract_merchant("UPI/DR/609178413960/SALEEM K/YESB/q763/UPI") == "Saleem K"
    # Axis P2A/P2M style
    assert extract_merchant("UPI/P2A/646193132733/XAVIER INFANTA MICHE /UPI/TAMILNAD MERCANTILE") == "Xavier Infanta Miche"
    assert extract_merchant("UPI/P2M/125430393560/Swiggy /NO REM/ICICI Bank") == "Swiggy"
    # Axis NEFT (non-greedy: must stop at first slash after name)
    assert extract_merchant("NEFT/BOFAH26097003281/ACCENTURE SOLUTIONS PVT LTD/BANK OF AMERICA/123") == "Accenture Solutions Pvt Ltd"
    # Axis mobile transfer
    assert extract_merchant("MOB/TPFT/THERESE SIRU MA/914010047242727") == "Therese Siru Ma"
    # Unknown format falls back to truncation, never crashes
    assert extract_merchant("") == ""
    assert len(extract_merchant("A" * 500)) <= 40


def test_anomaly_detection_flags_outliers():
    rows = [row(date(2026, 1, i + 1), debit=100.0) for i in range(10)]
    rows.append(row(date(2026, 1, 20), debit=50000.0))
    flagged = detect_anomalies(rows)
    assert len(flagged) == 1 and flagged[0]["debit"] == 50000.0


def test_anomaly_needs_min_samples():
    assert detect_anomalies([row(date(2026, 1, 1), debit=5.0)]) == []


def test_summaries_reconcile():
    rows = [
        row(date(2026, 1, 5), debit=100.0, merchant="A", category="Food"),
        row(date(2026, 1, 6), credit=500.0, merchant="B", category="Salary"),
        row(date(2026, 2, 1), debit=50.0, merchant="A", category="Food"),
    ]
    s = spending_stats(rows)
    assert s["total_spend"] == 150.0 and s["total_income"] == 500.0 and s["txn_count"] == 3
    m = {x["month"]: x for x in monthly_summary(rows)}
    assert m["Jan 2026"]["expense"] == 100.0 and m["Jan 2026"]["income"] == 500.0
    top = top_merchants(rows)
    assert top[0] == {"merchant": "A", "total_spend": 150.0}


def test_clean_and_enrich_dedupes_and_sorts():
    raw = [
        {"date": date(2026, 1, 2), "narration": "UPI-SWIGGY-SWIGGY@YBL-1234", "debit": 10.0, "credit": 0.0, "balance": 90.0},
        {"date": date(2026, 1, 2), "narration": "UPI-SWIGGY-SWIGGY@YBL-1234", "debit": 10.0, "credit": 0.0, "balance": 90.0},  # dupe
        {"date": date(2026, 1, 1), "narration": "UPI-ZOMATO-ZOMATO@PAYTM-5678", "debit": 5.0, "credit": 0.0, "balance": 100.0},
        {"date": None, "narration": "garbage", "debit": 1.0, "credit": 0.0, "balance": 0.0},  # dropped
    ]
    out = clean_and_enrich(raw)
    assert len(out) == 2
    assert out[0]["date"] < out[1]["date"]
    assert out[1]["merchant"] == "Swiggy"


def test_export_excel_produces_valid_workbook():
    rows = [
        row(date(2026, 1, 5), narration="UPI-SWIGGY-SWIGGY@YBL-1", debit=100.0, merchant="Swiggy", category="Food"),
        row(date(2026, 1, 6), narration="NEFT-SALARY", credit=50000.0, merchant="Employer", category="Salary"),
        row(date(2026, 1, 20), narration="UPI-BIGSPEND", debit=40000.0, merchant="BigSpend", category="Shopping"),
    ]
    anomalies = detect_anomalies(rows)
    for r in anomalies:
        r["is_anomaly"] = True
    monthly, cats, merchants = monthly_summary(rows), category_summary(rows), top_merchants(rows)
    stats = spending_stats(rows)

    buf = io.BytesIO()
    export_excel(rows, monthly, cats, merchants, anomalies, stats, buf)
    buf.seek(0)

    wb = load_workbook(buf)
    assert {"Summary", "Transactions", "Monthly Summary", "Categories", "Top Merchants", "Anomalies"} <= set(wb.sheetnames)
    assert wb["Transactions"].max_row >= len(rows) + 1  # +1 header row


def test_export_excel_handles_rows_without_ref_no():
    # Axis's extractor (_extract_axis) never sets ref_no/value_date, unlike
    # every other bank's extractor — export must not KeyError on that.
    rows = [{"date": date(2026, 1, 5), "narration": "NEFT/XYZ/PAYEE", "debit": 50.0,
             "credit": 0.0, "balance": 950.0, "merchant": "Payee", "category": "Other Expense",
             "is_anomaly": False}]
    monthly, cats, merchants = monthly_summary(rows), category_summary(rows), top_merchants(rows)
    stats = spending_stats(rows)

    buf = io.BytesIO()
    export_excel(rows, monthly, cats, merchants, [], stats, buf)  # must not raise
    buf.seek(0)
    wb = load_workbook(buf)
    assert wb["Transactions"].max_row == 2


def test_detect_bank_sbi_account_summary_layout_not_misdetected_as_hdfc():
    # Real-world bug: SBI's "Account Summary" layout renders its table header
    # ("Value Date", "Post Date", ...) as non-text, so _TABLE_START_MARKERS
    # never cuts off the header — the full text (including narrations) gets
    # scanned, and a UPI counterparty narration mentioning "HDFC" used to
    # false-positive-match before any SBI signature was found.
    text = """
    WELCOME MR SIVASUBRAMANIAN ACCOUNT SUMMARY
    CLEAR BALANCE : 2,827.68CR UNCLEARED AMOUNT : 0.00
    DRAWING POWER : 0.00
    01/04/2026 01/04/2026 UPI/DR/609255813673/MANOJBA/HDFC/mr.manojba/UPI 95.00 - 226.74
    """.upper()
    assert _match_bank_signature(text) == "SBI"


def test_detect_bank_still_finds_hdfc_when_genuinely_hdfc():
    text = "HDFC BANK STATEMENT OF ACCOUNT VALUE DATE NARRATION DEBIT CREDIT"
    assert _match_bank_signature(text) == "HDFC"


# ── fy_summary ────────────────────────────────────────────────────────────

def test_fy_summary_groups_april_to_march():
    rows = [
        row(date(2026, 3, 15), debit=100.0),   # FY 2025-26 (Jan-Mar belongs to prior FY)
        row(date(2026, 4, 1), debit=200.0),    # FY 2026-27 (April starts new FY)
        row(date(2027, 1, 10), debit=50.0),    # FY 2026-27
    ]
    fy = fy_summary(rows)
    by_fy = {y["fy"]: y for y in fy}
    assert by_fy["FY 2025-26"]["expense"] == 100.0
    assert by_fy["FY 2026-27"]["expense"] == 250.0


def test_fy_summary_sorted_chronologically():
    rows = [row(date(2027, 5, 1)), row(date(2025, 5, 1)), row(date(2026, 5, 1))]
    fy = fy_summary(rows)
    assert [y["fy"] for y in fy] == ["FY 2025-26", "FY 2026-27", "FY 2027-28"]


# ── Anonymized-export masking ────────────────────────────────────────────────

def test_mask_text_masks_long_digit_runs():
    assert _mask_text("ACH-DR-HDFC BANK LIMITED-0000161743135") == "ACH-DR-HDFC BANK LIMITED-*************"


def test_mask_text_masks_upi_handle_but_keeps_merchant():
    assert _mask_text("UPI-SWIGGY-SWIGGY@YBL-1234") == "UPI-SWIGGY-SW***@YBL-1234"


def test_mask_text_preserves_person_and_merchant_names():
    masked = _mask_text("UPI/DR/609178413960/SALEEM K/YESB/q763888035/UPI")
    assert "SALEEM K" in masked
    assert "609178413960" not in masked


def test_mask_ref_keeps_last_three_chars():
    assert _mask_ref("0000642625972044") == "XXXXXXXXXXXXX044"


def test_mask_ref_short_ref_fully_masked():
    assert _mask_ref("12") == "**"
    assert _mask_ref("") == ""


# ── New Excel export features ────────────────────────────────────────────────

def _export(rows_, masked=False):
    monthly, cats, merchants = monthly_summary(rows_), category_summary(rows_), top_merchants(rows_)
    anomalies = detect_anomalies(rows_)
    for r in anomalies:
        r["is_anomaly"] = True
    stats = spending_stats(rows_)
    buf = io.BytesIO()
    export_excel(rows_, monthly, cats, merchants, anomalies, stats, buf, masked=masked)
    buf.seek(0)
    return load_workbook(buf)


def test_transactions_sheet_has_tax_dropdown_column_and_validation():
    rows = [row(date(2026, 1, i + 1), debit=100.0) for i in range(3)]
    wb = _export(rows)
    ws = wb["Transactions"]
    assert ws.cell(row=1, column=11).value == "Tax Deductible?"
    assert len(ws.data_validations.dataValidation) == 1
    dv = ws.data_validations.dataValidation[0]
    assert "Yes - Business" in dv.formula1


def test_summary_sheet_has_tax_sumif_formula():
    rows = [row(date(2026, 1, i + 1), debit=100.0) for i in range(3)]
    wb = _export(rows)
    ws = wb["Summary"]
    formulas = [c.value for r in ws.iter_rows() for c in r if isinstance(c.value, str) and c.value.startswith("=SUMIF")]
    assert any("Transactions!K:K" in f and "Transactions!F:F" in f for f in formulas)


def test_summary_sheet_has_fy_breakdown():
    rows = [row(date(2026, 1, 5), debit=100.0), row(date(2026, 5, 5), debit=200.0)]
    wb = _export(rows)
    ws = wb["Summary"]
    values = [c.value for r in ws.iter_rows() for c in r]
    assert "FY 2025-26" in values and "FY 2026-27" in values


def test_anomaly_hyperlink_targets_correct_transaction_row():
    rows = [row(date(2026, 1, i + 1), debit=100.0) for i in range(10)]
    rows.append(row(date(2026, 1, 20), debit=50000.0))  # the anomaly, last row -> row 12 in sheet
    wb = _export(rows)
    ws = wb["Anomalies"]
    jump_cell = ws.cell(row=8, column=6)  # first (only) anomaly data row
    assert jump_cell.hyperlink is not None
    assert jump_cell.hyperlink.target == "#'Transactions'!A12"


def test_masked_export_scrubs_narration_and_ref():
    rows = [{"date": date(2026, 1, 5), "narration": "UPI-SWIGGY-SWIGGY@YBL-1234", "debit": 100.0,
             "credit": 0.0, "balance": 900.0, "merchant": "Swiggy", "category": "Food & Dining",
             "is_anomaly": False, "ref_no": "0000642625972044", "value_date": date(2026, 1, 5)}]
    wb = _export(rows, masked=True)
    ws = wb["Transactions"]
    assert ws.cell(row=2, column=3).value == "UPI-SWIGGY-SW***@YBL-1234"
    assert ws.cell(row=2, column=4).value == "XXXXXXXXXXXXX044"


def test_unmasked_export_keeps_narration_untouched():
    rows = [{"date": date(2026, 1, 5), "narration": "UPI-SWIGGY-SWIGGY@YBL-1234", "debit": 100.0,
             "credit": 0.0, "balance": 900.0, "merchant": "Swiggy", "category": "Food & Dining",
             "is_anomaly": False, "ref_no": "0000642625972044", "value_date": date(2026, 1, 5)}]
    wb = _export(rows, masked=False)
    ws = wb["Transactions"]
    assert ws.cell(row=2, column=3).value == "UPI-SWIGGY-SWIGGY@YBL-1234"


def test_indian_currency_format_applied_to_amount_columns():
    rows = [row(date(2026, 1, 5), debit=100.0, credit=0.0)]
    wb = _export(rows)
    ws = wb["Transactions"]
    debit_cell = ws.cell(row=2, column=6)
    assert ">=10000000" in debit_cell.number_format and ">=100000" in debit_cell.number_format


def test_freeze_panes_set_on_every_sheet():
    rows = [row(date(2026, 1, 5), debit=100.0)]
    wb = _export(rows)
    for name in ["Transactions", "Monthly Summary", "Categories", "Top Merchants", "Anomalies"]:
        assert wb[name].freeze_panes is not None, f"{name} has no freeze_panes set"


# ── IOB word-based parser — previously zero coverage ─────────────────────
# Fake pdfplumber page/pdf: _extract_iob only needs pdf.pages[i].extract_words().
class _FakeIOBPage:
    def __init__(self, words):
        self._words = words

    def extract_words(self, x_tolerance=3, y_tolerance=3):
        return self._words


class _FakeIOBPDF:
    def __init__(self, pages):
        self.pages = pages


def _iob_word(text, x0, top):
    return {"text": text, "x0": x0, "top": top}


def _iob_header():
    return [_iob_word("Date(Value", 49, 205)]


def test_iob_short_narration_parses_normally():
    words = _iob_header() + [
        _iob_word("01-Nov-25", 49, 225),
        _iob_word("UPI-SWIGGY-abc", 115, 225),
        _iob_word("500.00", 415, 225),
    ]
    rows = _extract_iob(_FakeIOBPDF([_FakeIOBPage(words)]))
    assert len(rows) == 1
    assert rows[0]["narration"] == "UPI-SWIGGY-abc"
    assert rows[0]["debit"] == 500.0


def test_iob_long_narration_no_longer_truncated_at_ref_column():
    """Regression test for a real bug: narration text overflowing past the
    fixed-width narration column (x < 278) into the neighbouring ref-number
    column's x-range used to be silently dropped, truncating narrations
    like 'Pay on UPI/merchant/xyz long description' down to just 'Pay on
    UPI/merchant/xyz' — the exact shape of a real user-reported issue."""
    words = _iob_header() + [
        _iob_word("05-Nov-25", 49, 250),
        _iob_word("Pay", 115, 250),
        _iob_word("on", 140, 250),
        _iob_word("UPI/merchant/xyz", 200, 250),
        _iob_word("long", 285, 250),          # spills past narration's x1=278
        _iob_word("description", 320, 250),   # spills further
        _iob_word("900.00", 415, 250),
    ]
    rows = _extract_iob(_FakeIOBPDF([_FakeIOBPage(words)]))
    assert len(rows) == 1
    assert rows[0]["narration"] == "Pay on UPI/merchant/xyz long description"
    assert rows[0]["ref_no"] == ""


def test_iob_real_reference_code_still_goes_to_ref_not_narration():
    words = _iob_header() + [
        _iob_word("06-Nov-25", 49, 270),
        _iob_word("NEFT-payment", 115, 270),
        _iob_word("IOBR52401234567", 285, 270),
        _iob_word("300.00", 415, 270),
    ]
    rows = _extract_iob(_FakeIOBPDF([_FakeIOBPage(words)]))
    assert len(rows) == 1
    assert rows[0]["narration"] == "NEFT-payment"
    assert rows[0]["ref_no"] == "IOBR52401234567"
